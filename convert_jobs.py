import json
import openpyxl

wb = openpyxl.load_workbook("jobs excerise.xlsx")
ws = wb.active

headers = [cell.value for cell in ws[1]]
col = {name: idx for idx, name in enumerate(headers)}

jobs = []
for row in ws.iter_rows(min_row=2):
    if all(cell.value is None for cell in row):
        continue

    # At-risk rows have a solid theme fill with tint ≈ -0.5 on the first cell
    at_risk = False
    try:
        f = row[0].fill
        if f.fill_type == 'solid' and abs(f.fgColor.tint - (-0.5)) < 0.01:
            at_risk = True
    except Exception:
        pass

    jobs.append({
        "career_cluster": row[col["Career Cluster"]].value,
        "career_pathway": row[col["Career Pathway"]].value,
        "code": row[col["Code"]].value,
        "occupation": row[col["Occupation"]].value,
        "description": row[col["Description"]].value,
        "at_risk": at_risk,
    })

with open("public/jobs.json", "w", encoding="utf-8") as f:
    json.dump(jobs, f, indent=2, ensure_ascii=False)

at_risk_count = sum(1 for j in jobs if j["at_risk"])
print(f"Wrote {len(jobs)} jobs ({at_risk_count} at-risk) to public/jobs.json")
